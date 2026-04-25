"""
MatrixFileHandler - Persistencia de Matrices en Formato Excel

Manejo de lectura y escritura de matrices de multiplicación en archivos Excel (.xlsx).
Este módulo reemplaza la persistencia XML anterior.

Formato del archivo Excel:
    - Hoja "Matriz A": Datos de la primera matriz
    - Hoja "Matriz B": Datos de la segunda matriz
    - Hoja "Info": Metadatos (caso, tamaño, fecha)

Funcionalidad:
    - Guardar matrices en formato Excel para legibilidad
    - Cargar matrices desde archivos Excel
    - Incluir metadatos del caso de prueba
"""

import os
from pathlib import Path
from datetime import datetime
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class MatrixFileHandler:
    """
    Manejador de persistencia de matrices en formato Excel.

    Attributes:
        DEFAULT_DIRECTORY (Path): Ruta por defecto para guardar archivos Excel
    """

    DEFAULT_DIRECTORY = Path(__file__).resolve().parent.parent.parent / "src/main/resources/matrices"

    @staticmethod
    def save_matrix(matrix_a, matrix_b, size, case):
        """
        Guarda dos matrices en un archivo Excel.

        Args:
            matrix_a (np.ndarray or list): Primera matriz (N×N)
            matrix_b (np.ndarray or list): Segunda matriz (N×N)
            size (int): Dimensión de las matrices (N)
            case (str): Identificador del caso ("Caso1" o "Caso2")

        Returns:
            str: Ruta del archivo Excel creado

        Raises:
            OSError: Si no se puede crear el directorio o archivo
        """
        if not MatrixFileHandler.DEFAULT_DIRECTORY.exists():
            os.makedirs(MatrixFileHandler.DEFAULT_DIRECTORY)

        filename = f"matrix_{case}_{size}x{size}.xlsx"
        filepath = os.path.join(MatrixFileHandler.DEFAULT_DIRECTORY, filename)

        wb = Workbook()

        MatrixFileHandler._create_matrix_sheet(wb, "Matriz A", matrix_a, size)
        MatrixFileHandler._create_matrix_sheet(wb, "Matriz B", matrix_b, size)
        MatrixFileHandler._create_info_sheet(wb, size, case)

        wb.save(filepath)
        return filepath

    @staticmethod
    def _create_matrix_sheet(workbook, sheet_name, matrix, size):
        """
        Crea una hoja de Excel con datos de matriz.

        Args:
            workbook (Workbook): Workbook de openpyxl
            sheet_name (str): Nombre de la hoja
            matrix (np.ndarray or list): Datos de la matriz
            size (int): Dimensión de la matriz
        """
        ws = workbook.create_sheet(sheet_name, 0)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws.cell(row=1, column=1, value=f"Matriz {sheet_name.split()[-1]} - {size}×{size}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(size, 10))

        for col in range(size):
            cell = ws.cell(row=3, column=col + 1, value=f"Col {col}")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for row in range(size):
            cell = ws.cell(row=row + 4, column=1, value=f"Fila {row}")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

            for col in range(size):
                value = matrix[row][col] if isinstance(matrix, list) else matrix[row, col]
                cell = ws.cell(row=row + 4, column=col + 2, value=float(value))
                cell.border = border
                cell.alignment = Alignment(horizontal="right")

        for col in range(size + 1):
            ws.column_dimensions[get_column_letter(col + 1)].width = 12

    @staticmethod
    def _create_info_sheet(workbook, size, case):
        """
        Crea una hoja de información con metadatos.

        Args:
            workbook (Workbook): Workbook de openpyxl
            size (int): Dimensión de las matrices
            case (str): Identificador del caso
        """
        ws = workbook.create_sheet("Info")
        ws.sheet_properties.tabColor = "70AD47"

        info_data = [
            ("Caso", case),
            ("Tamaño", f"{size}×{size}"),
            ("Fecha", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Elementos por matriz", size * size),
            ("Total matrices", 2),
            ("Formato numérico", "float64 (double)")
        ]

        for row, (label, value) in enumerate(info_data, start=2):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25

        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

    @staticmethod
    def load_matrix(size, case):
        """
        Carga matrices desde un archivo Excel.

        Args:
            size (int): Dimensión de las matrices
            case (str): Identificador del caso ("Caso1" o "Caso2")

        Returns:
            tuple: (matriz_a, matriz_b) ambas como np.ndarray

        Raises:
            FileNotFoundError: Si el archivo Excel no existe
        """
        filename = f"matrix_{case}_{size}x{size}.xlsx"
        filepath = os.path.join(MatrixFileHandler.DEFAULT_DIRECTORY, filename)

        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Archivo de matriz no encontrado: {filepath}")

        wb = load_workbook(filepath, data_only=True)

        ws_a = wb["Matriz A"]
        matrix_a = MatrixFileHandler._extract_matrix_from_sheet(ws_a, size)

        ws_b = wb["Matriz B"]
        matrix_b = MatrixFileHandler._extract_matrix_from_sheet(ws_b, size)

        return matrix_a, matrix_b

    @staticmethod
    def _extract_matrix_from_sheet(worksheet, size):
        """
        Extrae datos de matriz desde una hoja de Excel.

        Args:
            worksheet (Worksheet): Hoja de Excel con datos
            size (int): Dimensión esperada de la matriz

        Returns:
            np.ndarray: Matriz de dimensiones N×N
        """
        matrix = np.zeros((size, size), dtype=np.float64)

        for row in range(size):
            for col in range(size):
                cell = worksheet.cell(row=row + 4, column=col + 2)
                matrix[row, col] = float(cell.value) if cell.value is not None else 0.0

        return matrix
