"""
ResultsExcelHandler - Persistencia de Resultados y Gráficos en Excel

Manejo de lectura y escritura de resultados de tiempos de ejecución en archivos
Excel (.xlsx), incluyendo generación de gráficos comparativos.

Formato del archivo Excel:
    - Hoja "Caso1": Tiempos de ejecución para el primer caso configurado
    - Hoja "Caso2": Tiempos de ejecución para el segundo caso configurado
    - Hoja "Comparativa": Tabla resumen comparativa
    - Hoja "Gráfico": Gráfico comparativo embebido (imagen PNG)

Funcionalidad:
    - Guardar resultados de tiempos en formato Excel
    - Generar gráfico de barras comparativo de algoritmos
    - Exportar gráfico como imagen PNG
    - Insertar imagen del gráfico dentro del Excel
"""

import os
from pathlib import Path
from datetime import datetime
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
from matplotlib.patches import Patch

matplotlib.rcParams['font.family'] = 'DejaVu Sans'
matplotlib.rcParams['font.size'] = 10


class ResultsExcelHandler:
    """
    Manejador de persistencia de resultados de tiempos en formato Excel.

    Attributes:
        DEFAULT_DIRECTORY (Path): Ruta por defecto para guardar archivos Excel
        FILE_NAME (str): Nombre del archivo de resultados
    """

    DEFAULT_DIRECTORY = Path(__file__).resolve().parent.parent.parent / "src/main/resources/results"
    FILE_NAME = "python_results.xlsx"
    FILE_PATH = DEFAULT_DIRECTORY / FILE_NAME

    PNG_FILENAME = "grafico_comparativo.png"
    PNG_PATH = DEFAULT_DIRECTORY / PNG_FILENAME

    ALGORITHM_NAMES = [
        "NaivOnArray",
        "NaivLoopUnrollingTwo",
        "NaivLoopUnrollingFour",
        "WinogradOriginal",
        "WinogradScaled",
        "StrassenNaiv",
        "StrassenWinograd",
        "III_3_Sequential_Block",
        "III_4_Parallel_Block",
        "III_5_Enhanced_Parallel_Block",
        "IV_3_Sequential_Block",
        "IV_4_Parallel_Block",
        "IV_5_Enhanced_Parallel_Block",
        "V_3_Sequential_Block",
        "V_4_Parallel_Block"
    ]

    @staticmethod
    def _format_case_label(case_name, results):
        """
        Construye una etiqueta legible para un caso usando el tamano ejecutado.

        Args:
            case_name (str): Nombre del caso ("Caso1" o "Caso2")
            results (list): Resultados de ejecucion disponibles

        Returns:
            str: Etiqueta en formato "Caso 1 (128×128)"
        """
        base_label = case_name.replace("Caso", "Caso ")

        for result in results:
            if result.get("case") == case_name and result.get("size"):
                size = result["size"]
                return f"{base_label} ({size}×{size})"

        return base_label

    @staticmethod
    def save_result(size, algorithm, execution_time, case="", rows=0, cols=0, memory_kb=0, verified=False):
        """
        Guarda un resultado individual de ejecución.

        Args:
            size (int): Tamaño de la matriz
            algorithm (str): Nombre del algoritmo
            execution_time (int): Tiempo de ejecución en nanosegundos
            case (str): Identificador del caso ("Caso1" o "Caso2")
            rows (int): Filas de la matriz
            cols (int): Columnas de la matriz
            memory_kb (float): Pico de memoria en KB
            verified (bool): Indica si el resultado fue verificado como correcto
        """
        results = ResultsExcelHandler.load_results()
        result_data = {
            "size": size,
            "algorithm": algorithm,
            "language": "python",
            "executionTime": execution_time,
            "case": case,
            "rows": rows,
            "cols": cols,
            "memory_kb": memory_kb,
            "verified": verified
        }
        results.append(result_data)
        ResultsExcelHandler.save_all_results(results)

    @staticmethod
    def load_results():
        """
        Carga todos los resultados desde el archivo Excel.

        Returns:
            list: Lista de diccionarios con datos de resultados
        """
        if not ResultsExcelHandler.FILE_PATH.exists():
            return []
        try:
            wb = load_workbook(ResultsExcelHandler.FILE_PATH, data_only=True)
            results = []
            for sheet_name in ["Caso1", "Caso2"]:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for row in range(4, ws.max_row + 1):
                        algorithm = ws.cell(row=row, column=1).value
                        if algorithm and algorithm in ResultsExcelHandler.ALGORITHM_NAMES:
                            time_value = ws.cell(row=row, column=3).value
                            memory_value = ws.cell(row=row, column=5).value
                            verified_value = ws.cell(row=row, column=6).value
                            size_value = ws.cell(row=row, column=2).value
                            if time_value:
                                mem_kb = 0
                                if memory_value:
                                    mem_str = str(memory_value).replace(' KB', '').replace(' MB', '').strip()
                                    try:
                                        mem_kb = float(mem_str)
                                        if 'MB' in str(memory_value):
                                            mem_kb *= 1024
                                    except ValueError:
                                        mem_kb = 0
                                results.append({
                                    "size": int(size_value) if size_value else 0,
                                    "algorithm": algorithm,
                                    "language": "python",
                                    "executionTime": int(time_value),
                                    "case": sheet_name,
                                    "rows": int(size_value) if size_value else 0,
                                    "cols": int(size_value) if size_value else 0,
                                    "memory_kb": mem_kb,
                                    "verified": verified_value == "SI" if verified_value else False
                                })
            return results
        except Exception:
            return []

    @staticmethod
    def save_all_results(results):
        """
        Guarda todos los resultados en un archivo Excel.

        Args:
            results (list): Lista de diccionarios con datos de resultados
        """
        if not ResultsExcelHandler.DEFAULT_DIRECTORY.exists():
            os.makedirs(ResultsExcelHandler.DEFAULT_DIRECTORY)

        wb = Workbook()

        ResultsExcelHandler._create_case_sheet(wb, "Caso1", results)
        ResultsExcelHandler._create_case_sheet(wb, "Caso2", results)
        ResultsExcelHandler._create_summary_sheet(wb, results)

        wb.save(ResultsExcelHandler.FILE_PATH)

        ResultsExcelHandler._generate_and_save_chart(results)

    @staticmethod
    def _create_case_sheet(workbook, case_name, results):
        """
        Crea una hoja de Excel para un caso de prueba específico.

        Args:
            workbook: Workbook de openpyxl
            case_name (str): Nombre del caso ("Caso1" o "Caso2")
            results (list): Lista de resultados filtrados por caso
        """
        ws = workbook.create_sheet(case_name)
        ws.sheet_properties.tabColor = "4472C4" if case_name == "Caso1" else "ED7D31"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        title_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        case_label = ResultsExcelHandler._format_case_label(case_name, results)

        ws.cell(row=1, column=1, value=f"Resultados - {case_label}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).fill = title_fill
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

        ws.cell(row=2, column=1, value=f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=2, column=1).font = Font(italic=True, size=9)

        headers = ["Algoritmo", "Tamaño", "Tiempo (ns)", "Tiempo (ms)", "Memoria", "Verificado", "Caso"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        case_results = [r for r in results if r.get("case") == case_name]

        for row_idx, result in enumerate(case_results, start=4):
            ws.cell(row=row_idx, column=1, value=result["algorithm"]).border = border
            ws.cell(row=row_idx, column=2, value=result["size"]).border = border
            ws.cell(row=row_idx, column=3, value=result["executionTime"]).border = border
            ws.cell(row=row_idx, column=4, value=round(result["executionTime"] / 1_000_000, 3)).border = border
            mem_kb = result.get("memory_kb", 0)
            mem_str = f"{mem_kb:.2f} KB" if mem_kb < 1024 else f"{mem_kb/1024:.2f} MB"
            ws.cell(row=row_idx, column=5, value=mem_str).border = border
            ws.cell(row=row_idx, column=6, value="SI" if result.get("verified", False) else "NO").border = border
            ws.cell(row=row_idx, column=7, value=result["case"]).border = border

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10

    @staticmethod
    def _create_summary_sheet(workbook, results):
        """
        Crea una hoja de resumen comparativo.

        Args:
            workbook: Workbook de openpyxl
            results (list): Lista de todos los resultados
        """
        ws = workbook.create_sheet("Comparativa")
        ws.sheet_properties.tabColor = "70AD47"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws.cell(row=1, column=1, value="Comparativa de Tiempos de Ejecución y Memoria")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

        headers = [
            "Algoritmo",
            "Caso 1 (ns)", "Caso 1 (ms)", "Caso 1 Mem",
            "Caso 2 (ns)", "Caso 2 (ms)", "Caso 2 Mem",
            "Verif. C1", "Verif. C2"
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        case1_results = {r["algorithm"]: r for r in results if r.get("case") == "Caso1"}
        case2_results = {r["algorithm"]: r for r in results if r.get("case") == "Caso2"}

        for row_idx, alg_name in enumerate(ResultsExcelHandler.ALGORITHM_NAMES, start=4):
            ws.cell(row=row_idx, column=1, value=alg_name).border = border

            res1 = case1_results.get(alg_name, {})
            res2 = case2_results.get(alg_name, {})

            time1 = res1.get("executionTime", 0)
            time2 = res2.get("executionTime", 0)
            mem1 = res1.get("memory_kb", 0)
            mem2 = res2.get("memory_kb", 0)
            verif1 = "SI" if res1.get("verified", False) else "NO"
            verif2 = "SI" if res2.get("verified", False) else "NO"

            ws.cell(row=row_idx, column=2, value=time1 if time1 else "").border = border
            ws.cell(row=row_idx, column=3, value=round(time1 / 1_000_000, 3) if time1 else "").border = border
            mem1_str = f"{mem1:.2f} KB" if mem1 < 1024 else f"{mem1/1024:.2f} MB"
            ws.cell(row=row_idx, column=4, value=mem1_str if mem1 else "").border = border
            ws.cell(row=row_idx, column=5, value=time2 if time2 else "").border = border
            ws.cell(row=row_idx, column=6, value=round(time2 / 1_000_000, 3) if time2 else "").border = border
            mem2_str = f"{mem2:.2f} KB" if mem2 < 1024 else f"{mem2/1024:.2f} MB"
            ws.cell(row=row_idx, column=7, value=mem2_str if mem2 else "").border = border
            ws.cell(row=row_idx, column=8, value=verif1).border = border
            ws.cell(row=row_idx, column=9, value=verif2).border = border

        ws.column_dimensions['A'].width = 28
        for col in range(2, 10):
            ws.column_dimensions[chr(64 + col)].width = 15

        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

    @staticmethod
    def _generate_and_save_chart(results):
        """
        Genera un gráfico de barras comparativo y lo guarda como PNG,
        luego lo inserta en el archivo Excel.

        Args:
            results (list): Lista de resultados
        """
        case1_results = {r["algorithm"]: r for r in results if r.get("case") == "Caso1"}
        case2_results = {r["algorithm"]: r for r in results if r.get("case") == "Caso2"}

        algorithms = ResultsExcelHandler.ALGORITHM_NAMES
        case1_label = ResultsExcelHandler._format_case_label("Caso1", results)
        case2_label = ResultsExcelHandler._format_case_label("Caso2", results)
        case1_times = [case1_results.get(alg, {}).get("executionTime", 0) / 1_000_000 for alg in algorithms]
        case2_times = [case2_results.get(alg, {}).get("executionTime", 0) / 1_000_000 for alg in algorithms]
        case1_mem = [case1_results.get(alg, {}).get("memory_kb", 0) / 1024 for alg in algorithms]
        case2_mem = [case2_results.get(alg, {}).get("memory_kb", 0) / 1024 for alg in algorithms]

        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(16, 14), dpi=100)

        x = np.arange(len(algorithms))
        width = 0.35

        bars1 = ax1.bar(x - width/2, case1_times, width, label=case1_label, color='#4472C4', edgecolor='black')
        bars2 = ax1.bar(x + width/2, case2_times, width, label=case2_label, color='#ED7D31', edgecolor='black')

        ax1.set_xlabel('Algoritmo', fontsize=12, fontweight='bold')
        ax1.set_ylabel('Tiempo de Ejecución (s)', fontsize=12, fontweight='bold')
        ax1.set_title('Comparación de Tiempos de Ejecución - Algoritmos de Multiplicación de Matrices', fontsize=14, fontweight='bold')
        ax1.set_xticks(x)
        ax1.set_xticklabels(algorithms, rotation=45, ha='right', fontsize=9)
        ax1.legend(fontsize=10)
        ax1.grid(True, axis='y', linestyle='--', alpha=0.7)

        def add_labels_time(bars):
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax1.annotate(f'{height:.1f}',
                               xy=(bar.get_x() + bar.get_width() / 2, height),
                               xytext=(0, 3),
                               textcoords="offset points",
                               ha='center', va='bottom', fontsize=6, rotation=45)

        add_labels_time(bars1)
        add_labels_time(bars2)

        bars3 = ax2.bar(x - width/2, case1_mem, width, label=case1_label, color='#70AD47', edgecolor='black')
        bars4 = ax2.bar(x + width/2, case2_mem, width, label=case2_label, color='#FF0000', edgecolor='black')

        ax2.set_xlabel('Algoritmo', fontsize=12, fontweight='bold')
        ax2.set_ylabel('Memoria Pico (MB)', fontsize=12, fontweight='bold')
        ax2.set_title('Comparación de Memoria Pico - Algoritmos de Multiplicación de Matrices', fontsize=14, fontweight='bold')
        ax2.set_xticks(x)
        ax2.set_xticklabels(algorithms, rotation=45, ha='right', fontsize=9)
        ax2.legend(fontsize=10)
        ax2.grid(True, axis='y', linestyle='--', alpha=0.7)

        def add_labels_mem(bars):
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax2.annotate(f'{height:.1f}',
                               xy=(bar.get_x() + bar.get_width() / 2, height),
                               xytext=(0, 3),
                               textcoords="offset points",
                               ha='center', va='bottom', fontsize=6, rotation=45)

        add_labels_mem(bars3)
        add_labels_mem(bars4)

        plt.tight_layout()
        plt.savefig(ResultsExcelHandler.PNG_PATH, bbox_inches='tight', facecolor='white')
        plt.close(fig)

        ResultsExcelHandler._insert_image_in_excel()

    @staticmethod
    def _insert_image_in_excel():
        """
        Inserta la imagen del gráfico PNG en el archivo Excel.
        """
        if not ResultsExcelHandler.FILE_PATH.exists():
            return

        if not os.path.exists(ResultsExcelHandler.PNG_PATH):
            return

        wb = load_workbook(ResultsExcelHandler.FILE_PATH)

        if "Gráfico" in wb.sheetnames:
            del wb["Gráfico"]

        ws = wb.create_sheet("Gráfico")
        ws.sheet_properties.tabColor = "FF0000"

        ws.cell(row=1, column=1, value="Gráfico Comparativo de Tiempos de Ejecución")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        try:
            img = XLImage(str(ResultsExcelHandler.PNG_PATH))
            img.width = 900
            img.height = 500
            ws.add_image(img, 'A3')
        except Exception as e:
            print(f"Error inserting image: {e}")

        wb.save(ResultsExcelHandler.FILE_PATH)

    @staticmethod
    def get_combined_results():
        """
        Retorna todos los resultados cargados (compatibilidad con código anterior).

        Returns:
            list: Lista de diccionarios con datos de resultados
        """
        return ResultsExcelHandler.load_results()
